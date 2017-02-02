# -*- coding: utf-8 -*-
# Copyright (c) 2015, Frappe Technologies and contributors
# For license information, please see license.txt

from __future__ import unicode_literals
import os, difflib
import frappe, json, csv
import frappe.async
import frappe.permissions
from frappe.model.document import Document

import time

from frappe import _

from frappe.utils.dateutils import parse_date
from frappe.utils import cint, cstr, flt, getdate, get_datetime
from frappe.utils import get_site_name, get_site_path, get_site_base_path, get_path
from frappe.utils.csvutils import read_csv_content_from_attached_file, getlink
from frappe.utils.data import format_datetime
from frappe.utils.background_jobs import enqueue

from openpyxl import load_workbook

def get_data_keys_definition():
		return frappe._dict({
			"data_separator": _('Start entering data below this line'),
			"main_table": _("Table") + ":",
			"parent_table": _("Parent Table") + ":",
			"columns": _("Column Name") + ":",
			"doctype": _("DocType") + ":"
		})

class DataImport(Document):

	def validate(self):
		'''
			Decide whether the file is templated
			Template data field can be ['no file', 'raw', 'template']
			If the template field is 'raw' then set preview data and selected columns
		'''
		if not frappe.flags.in_test:
			self.name = self.reference_doctype +" import on "+ format_datetime(self.creation)

		if not self.import_file:
			self.preview_data = None
			self.file_preview = None
			self.flag_file_preview = 0
			self.template = "no file"

		if self.import_file and self.template == "no file":
			file_path = os.getcwd()+get_site_path()[1:].encode('utf8')+self.import_file
			filename, file_extension = os.path.splitext(file_path)
			if file_extension == '.xlsx':	
				self.set_preview_data(file_path)
			elif file_extension == '.csv':
				self.template = "template"
				self.preview_data = ""
				frappe.msgprint(_("You have uploaded a template file"))
			else:
				frappe.throw("Unsupported file format")


	def set_preview_data(self, file_path):
		'''
			Store the first 25 rows or less in the preview data 
			and check whether the xlsx file is templated or not
		'''
		wb = load_workbook(filename=file_path)
		ws = wb.active

		if ws.cell(row=20, column=1).value == get_data_keys_definition().data_separator:
			self.template = "template"
			self.preview_data = ""
			frappe.msgprint(_("You have uploaded a template file"))

		else:
			frappe.msgprint(_("Map the columns of your file using the dropdown"))
			
			self.template = "raw"
			file_data = []
			max_randered_row = 25 if ws.max_row>25 else ws.max_row
			for row in ws.iter_rows(max_row=max_randered_row):
				tmp_list = []
				for cell in row:
					tmp_list.append(cell.value)
				if [x for x in tmp_list if x != None]:
					file_data.append(tmp_list)
			self.preview_data = json.dumps(file_data)

			# Considering user will upload same file in a particular doctype
			column_map = []
			if file_data and not self.selected_columns:  
				for cell in file_data[0]:
					column_map.append(self.get_matched_column(cell))
				self.selected_columns = json.dumps(column_map)


	def get_matched_column(self, column_name=None):
		'''
			Fuzzy match the fieldname and input column 
		'''
		new_doc = frappe.new_doc(self.reference_doctype)
		doc_field = [frappe.unscrub(field.fieldname) for field in new_doc.meta.fields if field.fieldtype not in 
					['Section Break','Column Break','HTML','Table','Button','Image','Fold','Heading']]
		
		max_match = 0
		matched_field = ''
		for field in doc_field:
			seq=difflib.SequenceMatcher(None, str(field), str(column_name))
			d=seq.ratio()*100
			if d > max_match:
				max_match = d
				matched_field = field
		
		if max_match > 70:
			return frappe.scrub(matched_field)
		else:
			return ""

	def file_import(self, file_path=None):
		'''
			Trigger on import button and push the task to the background worker
		'''	
		if not file_path:
			file_path = os.getcwd()+get_site_path()[1:].encode('utf8') + self.import_file
		filename, file_extension = os.path.splitext(file_path)

		if self.template == "raw":
			enqueue(import_raw, file_path=file_path, doc_name=self.name, job_name="data_import")

		elif self.template == "template":
			enqueue(import_template, doc_name=self.name, file_path=file_path, job_name="data_import")

		else:
			frappe.throw(_("Error - Upload file again"))


def import_raw(file_path,doc_name):
	'''
		Import the raw xlsx file. It will insert new records only.
	'''

	di_doc = frappe.get_doc("Data Import",doc_name)

	frappe.flags.mute_emails = di_doc.no_email

	if di_doc.submit_after_import and not cint(frappe.db.get_value("DocType",
			di_doc.reference_doctype, "is_submittable")):
		di_doc.submit_after_import = 0

	# check permissions
	if not frappe.permissions.can_import(di_doc.reference_doctype):
		frappe.flags.mute_emails = False
		return {"messages": [_("Not allowed to Import") + ": " + _(di_doc.reference_doctype)], "error": True}

	ret = []
	def log(msg):
		ret.append(msg)

	error = False
	column_map = json.loads(di_doc.selected_columns)
	
	mendatory_field_flag = False
	# mendatory_field_list = []
	
	for field in frappe.get_meta(di_doc.reference_doctype).fields:
		if field.reqd == 1 and field.fieldname not in column_map:
			# mendatory_field_list.append(field.fieldname)
			mendatory_field_flag = True
			
	if mendatory_field_flag:
		frappe.throw(_("Please select all mendatory fields"))

	wb = load_workbook(filename=file_path, read_only=True)
	ws = wb.active

	start =  int(di_doc.selected_row)
	for i,row in enumerate(ws.iter_rows(min_row=start)):

		if [x for x in row if x.value != None]:

			frappe.publish_realtime("data_import", {"progress": [i+1, ws.max_row]}, 
				user=frappe.session.user)

			try:
				new_doc = frappe.new_doc(di_doc.reference_doctype)
				for j,cell in enumerate(row):
					if column_map[j] and column_map[j] != "do_not_map":
						setattr(new_doc, column_map[j], cell.value)
				new_doc.insert()
				new_doc.save()
				log([i,(getlink(di_doc.reference_doctype,new_doc.name)),"Row Inserted"])

			except Exception, e:
				error = True
				if new_doc:
					frappe.errprint(new_doc if isinstance(new_doc, dict) else new_doc.as_dict())
				err_msg = frappe.local.message_log and "\n\n".join(frappe.local.message_log) or cstr(e)

				frappe.errprint(frappe.get_traceback())

				ret = []
				ret.append(err_msg)
				ret.append(frappe.get_traceback())

			finally:
				frappe.local.message_log = []

			if error:
				frappe.db.rollback()
			else:
				frappe.db.commit()

	if frappe.flags.in_test:
		return True
	else:
		log_message = {"messages": ret, "error": error}
		di_doc.log_details = json.dumps(log_message)
		di_doc.freeze_doctype = 1
		di_doc.save()
		frappe.db.commit()
	

def import_template(doc_name, file_path, rows = None, ignore_links=False, pre_process=None, via_console=False):
	"""upload the templated data"""

	filename, file_extension = os.path.splitext(file_path)
	di_doc = frappe.get_doc("Data Import",doc_name)	

	frappe.flags.in_import = True
	frappe.flags.mute_emails = di_doc.no_email

	def bad_template():
		frappe.throw(_("Please do not change the rows above {0}").format(get_data_keys_definition().data_separator))

	def get_start_row():
		for i, row in enumerate(rows[:50]):
			if row and row[0]==get_data_keys_definition().data_separator:
				return i+1
		bad_template()

	def get_header_row(key):
		return get_header_row_and_idx(key)[0]

	def get_header_row_and_idx(key):
		for i, row in enumerate(header):
			if row and row[0]==key:
				return row, i
		return [], -1

	def filter_empty_columns(columns):
		empty_cols = filter(lambda x: x in ("", None), columns)

		if empty_cols:
			if columns[-1*len(empty_cols):] == empty_cols:
				# filter empty columns if they exist at the end
				columns = columns[:-1*len(empty_cols)]
			else:
				frappe.msgprint(_("Please make sure that there are no empty columns in the file."),
					raise_exception=1)

		return columns

	def make_column_map():
		doctype_row, row_idx = get_header_row_and_idx(get_data_keys_definition().doctype)
		if row_idx == -1: # old style
			return

		dt = None
		for i, d in enumerate(doctype_row[1:]):
			if d not in ("~", "-"):
				if d and doctype_row[i] in (None, '' ,'~', '-', 'DocType:'):
					dt, parentfield = d, doctype_row[i+2] or None
					doctypes.append((dt, parentfield))
					column_idx_to_fieldname[(dt, parentfield)] = {}
					column_idx_to_fieldtype[(dt, parentfield)] = {}
				if dt:
					column_idx_to_fieldname[(dt, parentfield)][i+1] = rows[row_idx + 2][i+1]
					column_idx_to_fieldtype[(dt, parentfield)][i+1] = rows[row_idx + 4][i+1]

	def get_doc(start_idx):
		if doctypes:
			doc = {}
			for idx in xrange(start_idx, len(rows)):
				if (not doc) or main_doc_empty(rows[idx]):
					for dt, parentfield in doctypes:
						d = {}
						for column_idx in column_idx_to_fieldname[(dt, parentfield)]:
							try:
								fieldname = column_idx_to_fieldname[(dt, parentfield)][column_idx]
								fieldtype = column_idx_to_fieldtype[(dt, parentfield)][column_idx]

								d[fieldname] = rows[idx][column_idx]
								if fieldtype in ("Int", "Check"):
									d[fieldname] = cint(d[fieldname])
								elif fieldtype in ("Float", "Currency", "Percent"):
									d[fieldname] = flt(d[fieldname])
								elif fieldtype == "Date":
									d[fieldname] = getdate(parse_date(d[fieldname])) if d[fieldname] else None
								elif fieldtype == "Datetime":
									if d[fieldname]:
										if " " in d[fieldname]:
											_date, _time = d[fieldname].split()
										else:
											_date, _time = d[fieldname], '00:00:00'
										_date = parse_date(d[fieldname])
										d[fieldname] = get_datetime(_date + " " + _time)
									else:
										d[fieldname] = None
							except IndexError:
								pass

						# scrub quotes from name and modified
						if d.get("name") and d["name"].startswith('"'):
							d["name"] = d["name"][1:-1]

						if sum([0 if not val else 1 for val in d.values()]):
							d['doctype'] = dt
							if dt == doctype:
								doc.update(d)
							else:
								if not di_doc.only_new_records:
									d['parent'] = doc["name"]
								d['parenttype'] = doctype
								d['parentfield'] = parentfield
								doc.setdefault(d['parentfield'], []).append(d)
				else:
					break

			return doc
		else:
			doc = frappe._dict(zip(columns, rows[start_idx][1:]))
			doc['doctype'] = doctype
			return doc

	def main_doc_empty(row):
		return not (row and ((len(row) > 1 and row[1]) or (len(row) > 2 and row[2])))

	users = frappe.db.sql_list("select name from tabUser")
	def prepare_for_insert(doc):
		# don't block data import if user is not set
		# migrating from another system
		if not doc.owner in users:
			doc.owner = frappe.session.user
		if not doc.modified_by in users:
			doc.modified_by = frappe.session.user


	# header

	if not rows and file_extension == '.csv':
		rows = read_csv_content_from_attached_file(frappe.get_doc("Data Import", di_doc.name),
			di_doc.ignore_encoding_errors)

	if not rows and file_extension == ".xlsx":
		rows = []
		wb1 = load_workbook(filename=file_path, read_only=True)
		ws1 = wb1.active

		for row in ws1.iter_rows():
			tmp_list = []
			for cell in row:
				tmp_list.append(cell.value)
			rows.append(tmp_list)

	start_row = get_start_row()
	header = rows[:start_row]
	data = rows[start_row:]
	doctype = get_header_row(get_data_keys_definition().main_table)[1]
	columns = filter_empty_columns(get_header_row(get_data_keys_definition().columns)[1:])
	doctypes = []
	column_idx_to_fieldname = {}
	column_idx_to_fieldtype = {}

	if di_doc.submit_after_import and not cint(frappe.db.get_value("DocType",
			doctype, "is_submittable")):
		di_doc.submit_after_import = False

	parenttype = get_header_row(get_data_keys_definition().parent_table)

	if len(parenttype) > 1:
		parenttype = parenttype[1]

	# check permissions
	if not frappe.permissions.can_import(parenttype or doctype):
		frappe.flags.mute_emails = False
		return {"messages": [_("Not allowed to Import") + ": " + _(doctype)], "error": True}


	make_column_map()

	# delete child rows (if parenttype)
	parentfield = None
	if parenttype:
		parentfield = get_parent_field(doctype, parenttype)

		if di_doc.only_new_records:
			delete_child_rows(data, doctype)

	ret = []

	def log(msg):
		if via_console:
			print msg.encode('utf-8')
		else:
			ret.append(msg)

	def as_link(doctype, name):
		if via_console:
			return "{0}: {1}".format(doctype, name)
		else:
			return getlink(doctype, name)

	error = False
	total = len(data)
	for i, row in enumerate(data):
		# bypass empty rows
		if main_doc_empty(row):
			continue

		row_idx = i + start_row
		doc = None

		# publish task_update
		frappe.publish_realtime("data_import", {"progress": [i+1, total]},
			user=frappe.session.user)

		try:
			doc = get_doc(row_idx)
			if pre_process:
				pre_process(doc)

			if parentfield:
				parent = frappe.get_doc(parenttype, doc["parent"])
				doc = parent.append(parentfield, doc)
				parent.save()
				log([i, getlink(di_doc.reference_doctype,new_doc.name), "Row Inserted"])
			else:
				if di_doc.only_new_records and doc["name"] and frappe.db.exists(doctype, doc["name"]):
					original = frappe.get_doc(doctype, doc["name"])
					original_name = original.name
					original.update(doc)
					# preserve original name for case sensitivity
					original.name = original_name
					original.flags.ignore_links = ignore_links
					original.save()
					log([row_idx+1, getlink(original.doctype, original.name), "Row updated"])
					doc = original
				else:
					if not di_doc.only_update:
						doc = frappe.get_doc(doc)
						prepare_for_insert(doc)
						doc.flags.ignore_links = ignore_links
						doc.insert()
						log([row_idx+1, getlink(doc.doctype, doc.name), "Row inserted"])
					else:
						log([row_idx+1, "", "Row ignored"])
				if di_doc.submit_after_import:
					doc.submit()
					log([row_idx+1, getlink(doc.doctype, doc.name), "Row submitted"])

		except Exception, e:
			error = True
			if doc:
				frappe.errprint(doc if isinstance(doc, dict) else doc.as_dict())
			err_msg = frappe.local.message_log and "\n\n".join(frappe.local.message_log) or cstr(e)

			frappe.errprint(frappe.get_traceback())

			ret = []
			ret.append(err_msg)
			ret.append(frappe.get_traceback())

		finally:
			frappe.local.message_log = []

	if error:
		frappe.db.rollback()
	else:
		frappe.db.commit()

	frappe.flags.mute_emails = False
	frappe.flags.in_import = False

	if frappe.flags.in_test:
		return True
	else:
		log_message = {"messages": ret, "error": error}
		di_doc.log_details = json.dumps(log_message)
		di_doc.freeze_doctype = 1
		di_doc.save()
		frappe.db.commit()

	def get_parent_field(doctype, parenttype):
		parentfield = None

		# get parentfield
		if parenttype:
			for d in frappe.get_meta(parenttype).get_table_fields():
				if d.options==doctype:
					parentfield = d.fieldname
					break

			if not parentfield:
				frappe.msgprint(_("Did not find {0} for {0} ({1})").format("parentfield", parenttype, doctype))
				raise Exception

		return parentfield

	def delete_child_rows(rows, doctype):
		"""delete child rows for all parents"""
		for p in list(set([r[1] for r in rows])):
			if p:
				frappe.db.sql("""delete from `tab{0}` where parent=%s""".format(doctype), p)
