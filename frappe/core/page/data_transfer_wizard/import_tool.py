# Copyright (c) 2015, Frappe Technologies Pvt. Ltd. and Contributors
# MIT License. See license.txt

from __future__ import unicode_literals
import frappe

from openpyxl import load_workbook

from frappe.utils.xlsxutils import read_xlsx_content_from_uploaded_file


@frappe.whitelist()
def upload_data(show_data, column_map=None, row_no=None):
	print "============>>>>>>>>>>>>>> TESTING"
	xlsx_file_as_list = []
	
	print show_data;
	file_objct = read_xlsx_content_from_uploaded_file()
	
	wb = load_workbook(file_objct)
	ws = wb.active
		
	for row in ws.iter_rows():
		tmp_list = []
		for cell in row:
			tmp_list.append(cell.value)
		xlsx_file_as_list.append(tmp_list)

	if show_data == True:
		print "in show data"
		return xlsx_file_as_list[:21]
	else:
		print show_data, column_map,row_no
