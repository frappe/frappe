# Copyright (c) 2015, Frappe Technologies Pvt. Ltd. and Contributors
# License: MIT. See LICENSE
import datetime
import re
from io import BytesIO

import openpyxl
import xlrd
from openpyxl import load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook.child import INVALID_TITLE_REGEX

import frappe
<<<<<<< HEAD
=======
from frappe import _
from frappe.core.utils import html2text
from frappe.utils import cint
from frappe.utils.csvutils import FORMULA_TRIGGER_CHARS
>>>>>>> e07d907e43 (fix(xlsxutils): write formula-like strings as literal text in XLSX export)
from frappe.utils.html_utils import unescape_html

ILLEGAL_CHARACTERS_RE = re.compile(
	r"[\000-\010]|[\013-\014]|[\016-\037]|\uFEFF|\uFFFE|\uFFFF|[\uD800-\uDFFF]"
)


def get_excel_date_format():
	date_format = frappe.get_system_settings("date_format")
	time_format = frappe.get_system_settings("time_format")

	# Excel-compatible format
	date_format = date_format.replace("mm", "MM")

	return date_format, time_format


# return xlsx file object
def make_xlsx(data, sheet_name, wb=None, column_widths=None):
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook(write_only=True)

	sheet_name_sanitized = INVALID_TITLE_REGEX.sub(" ", sheet_name)
	ws = wb.create_sheet(sheet_name_sanitized, 0)

	for i, column_width in enumerate(column_widths):
		if column_width:
			ws.column_dimensions[get_column_letter(i + 1)].width = column_width

	row1 = ws.row_dimensions[1]
	row1.font = Font(name="Calibri", bold=True)

	date_format, time_format = get_excel_date_format()

	for row in data:
		clean_row = []
		for item in row:
			if isinstance(item, str) and (sheet_name not in ["Data Import Template", "Data Export"]):
				value = handle_html(item)
			else:
				value = item

			if isinstance(item, str) and next(ILLEGAL_CHARACTERS_RE.finditer(value), None):
				# Remove illegal characters from the string
				value = ILLEGAL_CHARACTERS_RE.sub("", value)

			if isinstance(value, datetime.date | datetime.datetime):
				number_format = date_format
				if isinstance(value, datetime.datetime):
					number_format = f"{date_format} {time_format}"

				cell = WriteOnlyCell(ws, value=value)
				cell.number_format = number_format
				clean_row.append(cell)
			else:
				clean_row.append(value)

		ws.append(clean_row)

<<<<<<< HEAD
	xlsx_file = BytesIO()
	wb.save(xlsx_file)
=======
	write = ws.write
	write_string = ws.write_string
	has_cell_formats = bool(cell_formats)
	get_cell_format = cell_formats.get

	for row_idx, row in enumerate(data):
		for col_idx, value in enumerate(row):
			is_formula_like = False

			if isinstance(value, str):
				if handle_html_content:
					value = handle_html(value)

				if illegal_chars_search(value):
					value = illegal_chars_sub("", value)

				is_formula_like = value.startswith(FORMULA_TRIGGER_CHARS)

			cell_format = get_cell_format((row_idx, col_idx)) if has_cell_formats else None

			if is_formula_like:
				# force literal text so the cell isn't parsed as a formula
				write_string(row_idx, col_idx, value, cell_format)
			else:
				write(row_idx, col_idx, value, cell_format)

	if not created_wb:
		return

	wb.close()
	xlsx_file.seek(0)
>>>>>>> e07d907e43 (fix(xlsxutils): write formula-like strings as literal text in XLSX export)
	return xlsx_file


def handle_html(data):
	from frappe.core.utils import html2text

	# return if no html tags found
	data = frappe.as_unicode(data)

	if "<" not in data or ">" not in data:
		return data

	h = unescape_html(data or "")

	try:
		value = html2text(h, strip_links=True, wrap=False)
	except Exception:
		# unable to parse html, send it raw
		return data

	value = ", ".join(value.split("  \n"))
	value = " ".join(value.split("\n"))
	return ", ".join(value.split("# "))


def read_xlsx_file_from_attached_file(file_url=None, fcontent=None, filepath=None, *, read_only=False):
	if file_url:
		_file = frappe.get_doc("File", {"file_url": file_url})
		filename = _file.get_full_path()
	elif fcontent:
		filename = BytesIO(fcontent)
	elif filepath:
		filename = filepath
	else:
		return

	rows = []
	wb1 = load_workbook(filename=filename, data_only=True, read_only=read_only)
	ws1 = wb1.active
	for row in ws1.iter_rows():
		rows.append([cell.value for cell in row])
	if read_only:
		wb1.close()
	return rows


def read_xls_file_from_attached_file(content):
	book = xlrd.open_workbook(file_contents=content)
	sheets = book.sheets()
	sheet = sheets[0]
	return [sheet.row_values(i) for i in range(sheet.nrows)]


def build_xlsx_response(data, filename):
	from frappe.desk.utils import provide_binary_file

	provide_binary_file(filename, "xlsx", make_xlsx(data, filename).getvalue())
